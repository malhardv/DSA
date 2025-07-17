import asyncio
import json
import logging
from typing import Any, Dict, List, Optional, Union
import openpyxl
from openpyxl import Workbook, load_workbook
from mcp.server import Server
from mcp.server.models import InitializationOptions
from mcp.server.stdio import stdio_server
from mcp.types import (
    Resource,
    Tool,
    TextContent,
    ImageContent,
    EmbeddedResource,
    LoggingLevel
)
import os
from pathlib import Path
import re

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelMCPServer:
    def __init__(self, excel_file_path: str):
        self.excel_file_path = excel_file_path
        self.workbook = None
        self.control_sheet = None
        self.cy_is_sheet = None
        self.ty_is_sheet = None
        
        # Define the field mappings for the Control sheet
        # Based on your image, mapping field names to their cell locations
        self.control_field_mapping = {
            # Base Overhaul section
            "labor_rate": "N10",
            "labor_rate_$/hr": "N10",
            "labor_hours": "O10",
            "labor_hours_hrs": "O10",
            
            # Material section
            "new_material": "N208",
            "new_material_cost": "N208",
            "used_material": "N209", 
            "used_material_cost": "N209",
            "local_purchase": "N210",
            "local_purchase_cost": "N210",
            "component_repair": "N211",
            "component_repair_cost": "N211",
            "outside_vendor_repair": "N212",
            "outside_vendor_repair_cost": "N212",
            "labor": "N213",
            "labor_cost": "N213",
            "other_to_date": "N214",
            "other_to_date_cost": "N214",
            
            # Options section
            "life": "N309",
            "life_years": "N309",
            "used_ltp": "N310",
            "used_ltp_cost": "N310",
            "test_cell": "N311",
            "test_cell_cost": "N311",
            "refurb_expansion": "N312",
            "refurb_expansion_cost": "N312",
            "on_wing_support": "N313",
            "on_wing_support_cost": "N313",
            "pcm_fcs": "N314",
            "pcm_fcs_cost": "N314",
            "mixed_fleet": "N315",
            "mixed_fleet_cost": "N315",
            "spare_engine_lease": "N316",
            "spare_engine_lease_cost": "N316",
            "transportation": "N317",
            "transportation_cost": "N317",
            
            # BSP section
            "bsp_to_date": "N496",
            "hi_pool_share": "N511",
            "sn_pool_share": "N512",
            "hi_equitable_share": "N536",
            "sn_equitable_share": "N551",
            "sn_hpc_booster_comps": "N561",
            
            # Other common fields that might be referenced
            "escalation": "O340",  # Assuming this exists
            "severity": "P340",    # Assuming this exists
        }
        
        self.load_workbook()
    
    def load_workbook(self):
        """Load the Excel workbook and get sheet references."""
        try:
            if os.path.exists(self.excel_file_path):
                self.workbook = load_workbook(self.excel_file_path)
                self.control_sheet = self.workbook.get_sheet_by_name('Control')
                self.cy_is_sheet = self.workbook.get_sheet_by_name('CY_IS')
                self.ty_is_sheet = self.workbook.get_sheet_by_name('TY_IS')
                logger.info(f"Loaded workbook: {self.excel_file_path}")
            else:
                logger.error(f"Excel file not found: {self.excel_file_path}")
                raise FileNotFoundError(f"Excel file not found: {self.excel_file_path}")
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            raise
    
    def save_workbook(self):
        """Save the workbook."""
        try:
            self.workbook.save(self.excel_file_path)
            logger.info(f"Saved workbook: {self.excel_file_path}")
        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            raise
    
    def find_field_cell(self, field_name: str) -> Optional[str]:
        """Find the cell location for a given field name."""
        # Normalize field name
        field_name_lower = field_name.lower().replace(' ', '_').replace('-', '_')
        
        # Direct mapping lookup
        if field_name_lower in self.control_field_mapping:
            return self.control_field_mapping[field_name_lower]
        
        # Fuzzy matching for similar field names
        for mapped_field, cell in self.control_field_mapping.items():
            if field_name_lower in mapped_field or mapped_field in field_name_lower:
                return cell
        
        return None
    
    def get_control_value(self, field_name: str) -> Any:
        """Get a value from the Control sheet."""
        cell_address = self.find_field_cell(field_name)
        if not cell_address:
            raise ValueError(f"Field '{field_name}' not found in control sheet mapping")
        
        try:
            cell_value = self.control_sheet[cell_address].value
            return cell_value
        except Exception as e:
            logger.error(f"Error reading cell {cell_address}: {e}")
            raise
    
    def set_control_value(self, field_name: str, value: Union[int, float, str]) -> str:
        """Set a value in the Control sheet."""
        cell_address = self.find_field_cell(field_name)
        if not cell_address:
            raise ValueError(f"Field '{field_name}' not found in control sheet mapping")
        
        try:
            # Convert value to appropriate type
            if isinstance(value, str):
                try:
                    # Try to convert to number if it looks like one
                    if '.' in value:
                        value = float(value)
                    else:
                        value = int(value)
                except ValueError:
                    # Keep as string if conversion fails
                    pass
            
            old_value = self.control_sheet[cell_address].value
            self.control_sheet[cell_address] = value
            self.save_workbook()
            
            return f"Updated {field_name} in cell {cell_address}: {old_value} -> {value}"
        except Exception as e:
            logger.error(f"Error setting cell {cell_address}: {e}")
            raise
    
    def get_sheet_data(self, sheet_name: str, range_spec: Optional[str] = None) -> List[List[Any]]:
        """Get data from a specified sheet."""
        try:
            if sheet_name == 'CY_IS':
                sheet = self.cy_is_sheet
            elif sheet_name == 'TY_IS':
                sheet = self.ty_is_sheet
            elif sheet_name == 'Control':
                sheet = self.control_sheet
            else:
                raise ValueError(f"Unknown sheet: {sheet_name}")
            
            if range_spec:
                # Get specific range
                cells = sheet[range_spec]
                if hasattr(cells, '__iter__') and not isinstance(cells, str):
                    # Multiple cells
                    data = []
                    for row in cells:
                        if hasattr(row, '__iter__'):
                            data.append([cell.value for cell in row])
                        else:
                            data.append([row.value])
                    return data
                else:
                    # Single cell
                    return [[cells.value]]
            else:
                # Get all data
                data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # Skip empty rows
                        data.append(list(row))
                return data
        except Exception as e:
            logger.error(f"Error reading sheet {sheet_name}: {e}")
            raise

# Initialize the MCP server
server = Server("excel-financial-model")

# You'll need to set this to your actual Excel file path
EXCEL_FILE_PATH = "path/to/your/financial_model.xlsx"  # Update this path
excel_server = None

@server.list_resources()
async def handle_list_resources() -> List[Resource]:
    """List available resources."""
    return [
        Resource(
            uri="excel://control",
            name="Control Sheet",
            description="Financial model control parameters",
            mimeType="application/json"
        ),
        Resource(
            uri="excel://cy_is",
            name="Current Year IS",
            description="Current year income statement data",
            mimeType="application/json"
        ),
        Resource(
            uri="excel://ty_is",
            name="Target Year IS", 
            description="Target year income statement data",
            mimeType="application/json"
        )
    ]

@server.read_resource()
async def handle_read_resource(uri: str) -> str:
    """Read a specific resource."""
    global excel_server
    if not excel_server:
        raise RuntimeError("Excel server not initialized")
    
    if uri == "excel://control":
        # Return control sheet field mappings
        return json.dumps({
            "available_fields": list(excel_server.control_field_mapping.keys()),
            "description": "Available fields in the Control sheet that can be updated"
        })
    elif uri == "excel://cy_is":
        data = excel_server.get_sheet_data("CY_IS")
        return json.dumps({"sheet_data": data})
    elif uri == "excel://ty_is":
        data = excel_server.get_sheet_data("TY_IS")
        return json.dumps({"sheet_data": data})
    else:
        raise ValueError(f"Unknown resource: {uri}")

@server.list_tools()
async def handle_list_tools() -> List[Tool]:
    """List available tools."""
    return [
        Tool(
            name="update_control_field",
            description="Update a field in the Control sheet of the financial model",
            inputSchema={
                "type": "object",
                "properties": {
                    "field_name": {
                        "type": "string",
                        "description": "Name of the field to update (e.g., 'labor_rate', 'new_material', 'life')"
                    },
                    "value": {
                        "type": ["number", "string"],
                        "description": "New value for the field"
                    }
                },
                "required": ["field_name", "value"]
            }
        ),
        Tool(
            name="get_control_field",
            description="Get the current value of a field in the Control sheet",
            inputSchema={
                "type": "object",
                "properties": {
                    "field_name": {
                        "type": "string",
                        "description": "Name of the field to retrieve"
                    }
                },
                "required": ["field_name"]
            }
        ),
        Tool(
            name="get_sheet_data",
            description="Get data from CY_IS or TY_IS sheets",
            inputSchema={
                "type": "object",
                "properties": {
                    "sheet_name": {
                        "type": "string",
                        "enum": ["CY_IS", "TY_IS"],
                        "description": "Name of the sheet to read"
                    },
                    "range_spec": {
                        "type": "string",
                        "description": "Optional Excel range specification (e.g., 'A1:C10')"
                    }
                },
                "required": ["sheet_name"]
            }
        ),
        Tool(
            name="list_available_fields",
            description="List all available fields that can be updated in the Control sheet",
            inputSchema={
                "type": "object",
                "properties": {},
                "required": []
            }
        )
    ]

@server.call_tool()
async def handle_call_tool(name: str, arguments: Dict[str, Any]) -> List[TextContent]:
    """Handle tool calls."""
    global excel_server
    if not excel_server:
        raise RuntimeError("Excel server not initialized")
    
    try:
        if name == "update_control_field":
            field_name = arguments["field_name"]
            value = arguments["value"]
            result = excel_server.set_control_value(field_name, value)
            return [TextContent(type="text", text=result)]
        
        elif name == "get_control_field":
            field_name = arguments["field_name"]
            value = excel_server.get_control_value(field_name)
            return [TextContent(type="text", text=f"{field_name}: {value}")]
        
        elif name == "get_sheet_data":
            sheet_name = arguments["sheet_name"]
            range_spec = arguments.get("range_spec")
            data = excel_server.get_sheet_data(sheet_name, range_spec)
            return [TextContent(type="text", text=json.dumps(data, indent=2))]
        
        elif name == "list_available_fields":
            fields = list(excel_server.control_field_mapping.keys())
            field_list = "\n".join([f"- {field}" for field in sorted(fields)])
            return [TextContent(type="text", text=f"Available fields:\n{field_list}")]
        
        else:
            raise ValueError(f"Unknown tool: {name}")
    
    except Exception as e:
        logger.error(f"Error in tool {name}: {e}")
        return [TextContent(type="text", text=f"Error: {str(e)}")]

async def main():
    """Main function to run the MCP server."""
    global excel_server, EXCEL_FILE_PATH
    
    # Update this path to your actual Excel file
    if EXCEL_FILE_PATH == "path/to/your/financial_model.xlsx":
        print("Please update EXCEL_FILE_PATH to point to your actual Excel file")
        return
    
    try:
        excel_server = ExcelMCPServer(EXCEL_FILE_PATH)
        logger.info("Excel MCP Server initialized successfully")
        
        async with stdio_server() as (read_stream, write_stream):
            await server.run(
                read_stream,
                write_stream,
                InitializationOptions(
                    server_name="excel-financial-model",
                    server_version="0.1.0",
                    capabilities=server.get_capabilities(
                        notification_options=None,
                        experimental_capabilities=None
                    )
                )
            )
    except Exception as e:
        logger.error(f"Error starting server: {e}")
        raise

if __name__ == "__main__":
    asyncio.run(main())